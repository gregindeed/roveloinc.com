from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── Shared styles ────────────────────────────────────────────────────────────
NAVY        = "1F3864"
MED_BLUE    = "2E5090"
LT_BLUE     = "D6E4F0"
GRAY        = "F2F2F2"
WHITE       = "FFFFFF"
YELLOW      = "FFFF00"

CURR_FMT    = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

def hdr_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def thin_border():
    s = Side(border_style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_hdr(cell, text, size=10, bold=True, color="FFFFFF", bg=NAVY):
    cell.value = text
    cell.font = Font(name="Arial", bold=bold, size=size, color=color)
    cell.fill = hdr_fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)

def apply_data(cell, value, color="000000", bold=False, num_fmt=None,
               bg=None, align="left"):
    cell.value = value
    cell.font = Font(name="Arial", size=10, color=color, bold=bold)
    if bg:
        cell.fill = hdr_fill(bg)
    if num_fmt:
        cell.number_format = num_fmt
    cell.alignment = Alignment(horizontal=align, vertical="center")

def set_col_widths(ws, widths_dict):
    for col, w in widths_dict.items():
        ws.column_dimensions[col].width = w


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – April Summary
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "April Summary"
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells("A1:C1")
ws1["A1"].value = "REYES TIRES, INC — APRIL 2026 FINANCIAL SUMMARY"
ws1["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws1["A1"].fill = hdr_fill(NAVY)
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 30

ws1.merge_cells("A2:C2")
ws1["A2"].value = "For the Month Ended April 30, 2026"
ws1["A2"].font = Font(name="Arial", italic=True, size=10, color="FFFFFF")
ws1["A2"].fill = hdr_fill(MED_BLUE)
ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 18

# Column headers
for col, txt in [("A", "Line Item"), ("B", "Amount"), ("C", "Notes")]:
    ws1[f"{col}3"].value = txt
    ws1[f"{col}3"].font = Font(name="Arial", bold=True, size=10,
                               color="FFFFFF")
    ws1[f"{col}3"].fill = hdr_fill(NAVY)
    ws1[f"{col}3"].alignment = Alignment(horizontal="center",
                                         vertical="center")
ws1.row_dimensions[3].height = 18

set_col_widths(ws1, {"A": 38, "B": 18, "C": 42})

# ── Revenue section ──
def sec_hdr(ws, row, label, bg=MED_BLUE):
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"].value = label
    ws[f"A{row}"].font = Font(name="Arial", bold=True, size=10,
                               color="FFFFFF")
    ws[f"A{row}"].fill = hdr_fill(bg)
    ws.row_dimensions[row].height = 16

def data_row(ws, row, label, val, is_formula=False, indent=2, note=""):
    ws[f"A{row}"].value = ("  " * indent) + label
    ws[f"A{row}"].font = Font(name="Arial", size=10, color="000000")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    if is_formula:
        ws[f"B{row}"].value = val
        ws[f"B{row}"].font = Font(name="Arial", size=10, bold=True,
                                   color="000000")
    else:
        ws[f"B{row}"].value = val
        ws[f"B{row}"].font = Font(name="Arial", size=10, color="0000FF")
    ws[f"B{row}"].number_format = CURR_FMT
    ws[f"B{row}"].alignment = Alignment(horizontal="right", vertical="center")
    # alternating row color
    bg = GRAY if row % 2 == 0 else WHITE
    for c in ["A", "B", "C"]:
        ws[f"{c}{row}"].fill = hdr_fill(bg)
    if note:
        ws[f"C{row}"].value = note
        ws[f"C{row}"].font = Font(name="Arial", size=9, italic=True,
                                   color="555555")
    ws.row_dimensions[row].height = 15

def total_row(ws, row, label, formula, indent=1):
    ws[f"A{row}"].value = ("  " * indent) + label
    ws[f"A{row}"].font = Font(name="Arial", bold=True, size=10,
                               color="000000")
    ws[f"A{row}"].fill = hdr_fill(LT_BLUE)
    ws[f"B{row}"].value = formula
    ws[f"B{row}"].font = Font(name="Arial", bold=True, size=10,
                               color="000000")
    ws[f"B{row}"].number_format = CURR_FMT
    ws[f"B{row}"].fill = hdr_fill(LT_BLUE)
    ws[f"B{row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"C{row}"].fill = hdr_fill(LT_BLUE)
    ws.row_dimensions[row].height = 16

# Row layout
sec_hdr(ws1, 4, "REVENUE")
data_row(ws1, 5,  "Card Sales (EPX)",         21933.52, note="EPX merchant settlements")
data_row(ws1, 6,  "Zelle Payments",            7090.00, note="Zelle received from customers")
data_row(ws1, 7,  "Check Deposits",            4480.00, note="Mobile check deposits")
data_row(ws1, 8,  "Other (refunds/returns)",   2920.99, note="Amazon refund + bounced check return + overdraft transfer")
total_row(ws1, 9, "TOTAL REVENUE", "=SUM(B5:B8)")

sec_hdr(ws1, 10, "COST OF GOODS SOLD")
data_row(ws1, 11, "Inventory - Tires",        10384.34, note="Tire purchases from vendors")
data_row(ws1, 12, "Inventory - Parts & Supplies", 3941.50, note="Parts, supplies, wholesale")
total_row(ws1, 13, "TOTAL COGS", "=SUM(B11:B12)")

# Gross profit
ws1.merge_cells("A14:C14")
ws1["A14"].value = ""
ws1.row_dimensions[14].height = 6

ws1[f"A15"].value = "GROSS PROFIT"
ws1[f"A15"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ws1[f"A15"].fill = hdr_fill(NAVY)
ws1[f"B15"].value = "=B9-B13"
ws1[f"B15"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ws1[f"B15"].number_format = CURR_FMT
ws1[f"B15"].fill = hdr_fill(NAVY)
ws1[f"B15"].alignment = Alignment(horizontal="right", vertical="center")
ws1[f"C15"].fill = hdr_fill(NAVY)
ws1.row_dimensions[15].height = 20

sec_hdr(ws1, 16, "OPERATING EXPENSES")
opex_items = [
    (17, "Payroll - Wages",            9881.60, "ADP Wage Pay (5 payrolls × $1,976.32)"),
    (18, "Payroll - Taxes",            2430.72, "ADP Tax deposits (4 × $607.68)"),
    (19, "Payroll - 401k",              150.00, "ADP 401k (5 × $30.00)"),
    (20, "Payroll - Workers Comp",      513.69, "ADP Pay-By-Pay insurance"),
    (21, "Payroll - Processing Fees",   562.00, "ADP processing fees"),
    (22, "Rent",                       3400.00, "Basil Hosmer — Check #4072"),
    (23, "Insurance - Business",       1138.82, "TAPCO Insurance"),
    (24, "Equipment Lease",             741.95, "TimePaymentcorp"),
    (25, "Vehicle - Loan Payment",      326.39, "Toyota Financial"),
    (26, "Vehicle - Repairs & Parts",   133.52, "BMW El Cajon"),
    (27, "Utilities - Phone",           244.89, "AT&T"),
    (28, "Utilities - Internet (CC)",   110.17, "Cox Communications — CC ...0214"),
    (29, "Security System (CC)",         98.08, "ADT Security — CC ...6514"),
    (30, "Merchant Processing Fees",     52.66, "EPX + FDMS fees"),
    (31, "Bank Fees",                    18.70, "Monthly + excess transaction fees"),
    (32, "Professional Services - Tax",  420.00, "A to Z Tax Firm — CC ...0214"),
    (33, "Labor / Subcontractors",       460.00, "GW labor + Zelle to Teo"),
    (34, "Licenses & Permits",          250.00, "Bureau of Auto Repair — Check #4071"),
    (35, "Transportation - Tolls",        2.50, "SANDAG FasTrak"),
    (36, "Meals & Entertainment",         95.37, "Gladly Coffee + Snapdragon Stadium"),
    (37, "Finance Charges (CC)",         211.76, "CC ...0214 $63.07 + ...0539 $51.72 + ...6514 $96.97"),
    (38, "Miscellaneous",               530.60, "Infinity Kat $400.60 + LGND SOCAL $130.00"),
    (39, "Owner Draw",                  700.00, "Francisco Reyes — Check #4092"),
]
for r, lbl, val, note in opex_items:
    data_row(ws1, r, lbl, val, note=note)

total_row(ws1, 40, "TOTAL OPERATING EXPENSES", "=SUM(B17:B39)")

# Spacer
ws1.merge_cells("A41:C41")
ws1["A41"].value = ""
ws1.row_dimensions[41].height = 6

# Net Income
ws1[f"A42"].value = "NET INCOME"
ws1[f"A42"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
ws1[f"A42"].fill = hdr_fill(NAVY)
ws1[f"B42"].value = "=B15-B40"
ws1[f"B42"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
ws1[f"B42"].number_format = CURR_FMT
ws1[f"B42"].fill = hdr_fill(NAVY)
ws1[f"B42"].alignment = Alignment(horizontal="right", vertical="center")
ws1[f"C42"].fill = hdr_fill(NAVY)
ws1.row_dimensions[42].height = 22

# Spacer
ws1.row_dimensions[43].height = 8

# Warning note
ws1.merge_cells("A44:C44")
ws1["A44"].value = ("⚠️ Personal charges on business CC (~$1,107.26) require "
                    "owner review and are NOT included in above figures.")
ws1["A44"].font = Font(name="Arial", size=9, bold=True, color="FF0000")
ws1["A44"].fill = hdr_fill("FFF3CD")
ws1["A44"].alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True)
ws1.row_dimensions[44].height = 28

# source note
ws1.merge_cells("A45:C45")
ws1["A45"].value = ("Source: Bank of America Checking Statement April 2026 + "
                    "BofA Visa Credit Card Statements April 2026")
ws1["A45"].font = Font(name="Arial", size=8, italic=True, color="555555")
ws1["A45"].alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[45].height = 14


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Checking - Deposits
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Checking - Deposits")
ws2.sheet_view.showGridLines = False

set_col_widths(ws2, {"A": 12, "B": 38, "C": 22, "D": 26, "E": 16,
                      "F": 4, "G": 4, "H": 28, "I": 16})

# Title
ws2.merge_cells("A1:E1")
ws2["A1"].value = "REYES TIRES, INC — CHECKING ACCOUNT DEPOSITS — APRIL 2026"
ws2["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
ws2["A1"].fill = hdr_fill(NAVY)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

# Column headers row 2
hdr_labels = ["Date", "Description", "Type", "Category", "Amount"]
for i, lbl in enumerate(hdr_labels, 1):
    c = ws2.cell(row=2, column=i)
    c.value = lbl
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = hdr_fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[2].height = 16

deposits = [
    # date_str, description, type_, category, amount
    ("04/01/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",         870.01),
    ("04/01/2026", "Zelle - Sandra Castillo",               "Zelle Received",   "Sales - Zelle Payments",        840.00),
    ("04/02/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",         730.01),
    ("04/02/2026", "Zelle - Julien Kenderson",              "Zelle Received",   "Sales - Zelle Payments",        300.00),
    ("04/02/2026", "Zelle - Cam DX Corp",                   "Zelle Received",   "Sales - Zelle Payments",        150.00),
    ("04/03/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",         715.02),
    ("04/03/2026", "Zelle - Zamora Valenzuela",             "Zelle Received",   "Sales - Zelle Payments",        240.00),
    ("04/06/2026", "Mobile Check Deposit",                  "Check Deposit",    "Sales - Check Payments",       3160.00),
    ("04/06/2026", "Zelle - Farhat Mojaddidi",              "Zelle Received",   "Sales - Zelle Payments",       2050.00),
    ("04/06/2026", "Zelle - Yusuf Anas",                    "Zelle Received",   "Sales - Zelle Payments",        870.00),
    ("04/06/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",         635.01),
    ("04/06/2026", "Zelle - Direct Autohaus",               "Zelle Received",   "Sales - Zelle Payments",        320.00),
    ("04/06/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",         180.00),
    ("04/07/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",         545.02),
    ("04/07/2026", "Zelle - Carlos Escobar Padron",         "Zelle Received",   "Sales - Zelle Payments",        240.00),
    ("04/08/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",          75.00),
    ("04/09/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",          75.00),
    ("04/10/2026", "Amazon Refund",                         "Refund",           "Other - Refund",                387.37),
    ("04/13/2026", "Mobile Check Deposit",                  "Check Deposit",    "Sales - Check Payments",       1320.00),
    ("04/13/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",         435.01),
    ("04/14/2026", "Return of Posted Check",                "Bank Return",      "Non-Operating - Bank Return",  1976.32),
    ("04/14/2026", "Overdraft Protection Transfer",         "Bank Transfer",    "Non-Operating - Transfer",      557.31),
    ("04/14/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",        4310.13),
    ("04/15/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",         815.02),
    ("04/15/2026", "Zelle - Cam DX Corp",                   "Zelle Received",   "Sales - Zelle Payments",        200.00),
    ("04/16/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",         420.01),
    ("04/17/2026", "Zelle - Marco Castrosanchez (VW Atlas oil change)", "Zelle Received", "Sales - Zelle Payments", 290.00),
    ("04/17/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",         280.00),
    ("04/20/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",        1570.04),
    ("04/20/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",         760.02),
    ("04/20/2026", "Zelle - Jorge Romero Pino",             "Zelle Received",   "Sales - Zelle Payments",        280.00),
    ("04/20/2026", "Zelle - Faisal Anas",                   "Zelle Received",   "Sales - Zelle Payments",         70.00),
    ("04/20/2026", "Zelle - Carlos Calderon",               "Zelle Received",   "Sales - Zelle Payments",         60.00),
    ("04/21/2026", "Zelle - Direct Autohaus (BMW/Audi sensors)", "Zelle Received", "Sales - Zelle Payments",    350.00),
    ("04/21/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",          60.00),
    ("04/21/2026", "Zelle - Xdrive LLC",                    "Zelle Received",   "Sales - Zelle Payments",         50.00),
    ("04/22/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",         775.01),
    ("04/23/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",         795.02),
    ("04/24/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",        1000.01),
    ("04/24/2026", "Zelle - Cam DX Corp",                   "Zelle Received",   "Sales - Zelle Payments",        100.00),
    ("04/24/2026", "Zelle - Omar Zurita",                   "Zelle Received",   "Sales - Zelle Payments",         50.00),
    ("04/27/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",        1950.06),
    ("04/27/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",        1433.03),
    ("04/27/2026", "Zelle - Victor Villegas",               "Zelle Received",   "Sales - Zelle Payments",        100.00),
    ("04/28/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",        1230.03),
    ("04/29/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",        1215.04),
    ("04/29/2026", "Zelle - Jorge Ruiz (Lincoln)",          "Zelle Received",   "Sales - Zelle Payments",        200.00),
    ("04/29/2026", "Zelle - Richard Prieto (tire patch)",   "Zelle Received",   "Sales - Zelle Payments",         70.00),
    ("04/29/2026", "Zelle - Julien Kenderson",              "Zelle Received",   "Sales - Zelle Payments",         60.00),
    ("04/30/2026", "EPX Merchant Settlement",              "EPX Settlement",   "Sales - Card Payments",        1060.02),
    ("04/30/2026", "Zelle - Motor Point LLC",               "Zelle Received",   "Sales - Zelle Payments",        200.00),
]

start_row = 3
for i, (dt, desc, typ, cat, amt) in enumerate(deposits):
    r = start_row + i
    bg = GRAY if i % 2 == 0 else WHITE
    ws2.cell(row=r, column=1).value = dt
    ws2.cell(row=r, column=1).font = Font(name="Arial", size=10)
    ws2.cell(row=r, column=1).fill = hdr_fill(bg)
    ws2.cell(row=r, column=2).value = desc
    ws2.cell(row=r, column=2).font = Font(name="Arial", size=10)
    ws2.cell(row=r, column=2).fill = hdr_fill(bg)
    ws2.cell(row=r, column=3).value = typ
    ws2.cell(row=r, column=3).font = Font(name="Arial", size=10)
    ws2.cell(row=r, column=3).fill = hdr_fill(bg)
    ws2.cell(row=r, column=4).value = cat
    ws2.cell(row=r, column=4).font = Font(name="Arial", size=10)
    ws2.cell(row=r, column=4).fill = hdr_fill(bg)
    ws2.cell(row=r, column=5).value = amt
    ws2.cell(row=r, column=5).font = Font(name="Arial", size=10,
                                          color="0000FF")
    ws2.cell(row=r, column=5).number_format = CURR_FMT
    ws2.cell(row=r, column=5).fill = hdr_fill(bg)
    ws2.cell(row=r, column=5).alignment = Alignment(horizontal="right")
    ws2.row_dimensions[r].height = 15

# Total row
total_r = start_row + len(deposits)
ws2.cell(row=total_r, column=1).value = "TOTAL"
ws2.cell(row=total_r, column=1).font = Font(name="Arial", bold=True, size=10,
                                             color="000000")
ws2.cell(row=total_r, column=1).fill = hdr_fill(LT_BLUE)
for col in range(2, 6):
    ws2.cell(row=total_r, column=col).fill = hdr_fill(LT_BLUE)
ws2.cell(row=total_r, column=5).value = f"=SUM(E{start_row}:E{total_r-1})"
ws2.cell(row=total_r, column=5).font = Font(name="Arial", bold=True, size=10)
ws2.cell(row=total_r, column=5).number_format = CURR_FMT
ws2.cell(row=total_r, column=5).fill = hdr_fill(LT_BLUE)
ws2.cell(row=total_r, column=5).alignment = Alignment(horizontal="right")
ws2.row_dimensions[total_r].height = 16

# Category breakdown table (cols H-I)
cat_hdr_row = 2
ws2.cell(row=cat_hdr_row, column=8).value = "Category"
ws2.cell(row=cat_hdr_row, column=8).font = Font(name="Arial", bold=True,
                                                  size=10, color="FFFFFF")
ws2.cell(row=cat_hdr_row, column=8).fill = hdr_fill(NAVY)
ws2.cell(row=cat_hdr_row, column=9).value = "Total"
ws2.cell(row=cat_hdr_row, column=9).font = Font(name="Arial", bold=True,
                                                  size=10, color="FFFFFF")
ws2.cell(row=cat_hdr_row, column=9).fill = hdr_fill(NAVY)

d_categories = [
    "Sales - Card Payments",
    "Sales - Zelle Payments",
    "Sales - Check Payments",
    "Non-Operating - Bank Return",
    "Non-Operating - Transfer",
    "Other - Refund",
]
for j, cat in enumerate(d_categories):
    br = cat_hdr_row + 1 + j
    ws2.cell(row=br, column=8).value = cat
    ws2.cell(row=br, column=8).font = Font(name="Arial", size=10)
    ws2.cell(row=br, column=8).fill = hdr_fill(GRAY if j % 2 == 0 else WHITE)
    formula = (f'=SUMIF($D${start_row}:$D${total_r-1},'
               f'H{br},$E${start_row}:$E${total_r-1})')
    ws2.cell(row=br, column=9).value = formula
    ws2.cell(row=br, column=9).font = Font(name="Arial", size=10)
    ws2.cell(row=br, column=9).number_format = CURR_FMT
    ws2.cell(row=br, column=9).fill = hdr_fill(GRAY if j % 2 == 0 else WHITE)
    ws2.cell(row=br, column=9).alignment = Alignment(horizontal="right")

# Grand total for breakdown
gt_r = cat_hdr_row + 1 + len(d_categories)
ws2.cell(row=gt_r, column=8).value = "GRAND TOTAL"
ws2.cell(row=gt_r, column=8).font = Font(name="Arial", bold=True, size=10)
ws2.cell(row=gt_r, column=8).fill = hdr_fill(LT_BLUE)
ws2.cell(row=gt_r, column=9).value = (f"=SUM(I{cat_hdr_row+1}:I{gt_r-1})")
ws2.cell(row=gt_r, column=9).font = Font(name="Arial", bold=True, size=10)
ws2.cell(row=gt_r, column=9).number_format = CURR_FMT
ws2.cell(row=gt_r, column=9).fill = hdr_fill(LT_BLUE)
ws2.cell(row=gt_r, column=9).alignment = Alignment(horizontal="right")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – Checking - Expenses
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Checking - Expenses")
ws3.sheet_view.showGridLines = False

set_col_widths(ws3, {"A": 12, "B": 9, "C": 36, "D": 26, "E": 16,
                      "F": 4, "G": 4, "H": 4, "I": 30, "J": 16})

ws3.merge_cells("A1:E1")
ws3["A1"].value = "REYES TIRES, INC — CHECKING ACCOUNT EXPENSES — APRIL 2026"
ws3["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
ws3["A1"].fill = hdr_fill(NAVY)
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 26

exp_hdrs = ["Date", "Check #", "Description", "Category", "Amount"]
for i, lbl in enumerate(exp_hdrs, 1):
    c = ws3.cell(row=2, column=i)
    c.value = lbl
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = hdr_fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[2].height = 16

expenses = [
    # date, check_num, description, category, amount
    ("04/01/2026", "",      "Bank Monthly Maintenance Fee",           "Bank Fees",                   16.00),
    ("04/01/2026", "",      "Excess Transaction Fee",                  "Bank Fees",                    2.70),
    ("04/01/2026", "",      "EPX Merchant Processing Fee",             "Merchant Processing",         10.00),
    ("04/01/2026", "",      "ADP 401k Contribution",                   "Payroll-401k",                30.00),
    ("04/01/2026", "",      "ADP Pay-By-Pay (Workers Comp)",           "Payroll-Workers Comp",        86.75),
    ("04/01/2026", "4064",  "Tire Network — Tire Purchase",            "Inventory-Tires",            600.00),
    ("04/01/2026", "4066",  "U.S. Auto Force — Parts",                 "Inventory-Parts & Supplies", 574.38),
    ("04/01/2026", "4068",  "Tire Co — Tire Purchase",                 "Inventory-Tires",            200.00),
    ("04/03/2026", "",      "ADP Payroll Processing Fee",              "Payroll-Processing Fees",    118.50),
    ("04/03/2026", "",      "ADP Payroll Processing Fee",              "Payroll-Processing Fees",     40.00),
    ("04/03/2026", "",      "Toyota Financial — Vehicle Loan",         "Vehicle-Loan",               326.39),
    ("04/03/2026", "4069",  "Tire Outlet — Tire Purchase",             "Inventory-Tires",            630.00),
    ("04/06/2026", "",      "ADP Wage Pay — Payroll",                  "Payroll-Wages",             1976.32),
    ("04/06/2026", "",      "ADP Tax Deposit",                         "Payroll-Payroll Taxes",      607.68),
    ("04/06/2026", "",      "Zelle to Teo — Labor/Supply",             "Labor/Subcontractors",       200.00),
    ("04/06/2026", "",      "Zelle to Juan Used Tires",                "Inventory-Tires",            700.00),
    ("04/06/2026", "",      "Gladly Coffee",                           "Meals/Entertainment",         27.10),
    ("04/06/2026", "4070",  "Freedom — Parts Purchase",                "Inventory-Parts & Supplies", 479.40),
    ("04/07/2026", "",      "ADP 401k Contribution",                   "Payroll-401k",                30.00),
    ("04/07/2026", "",      "ADP Pay-By-Pay (Workers Comp)",           "Payroll-Workers Comp",       166.69),
    ("04/07/2026", "",      "FDMS Merchant Processing Fee",            "Merchant Processing",          4.95),
    ("04/07/2026", "",      "LGND SOCAL",                              "Miscellaneous",              130.00),
    ("04/08/2026", "",      "CC Payment — Card 6514",                  "Credit Card Payment",       1000.00),
    ("04/08/2026", "",      "FDMS Merchant Processing Fee",            "Merchant Processing",         37.71),
    ("04/10/2026", "",      "ADP Payroll Processing Fee",              "Payroll-Processing Fees",    134.50),
    ("04/10/2026", "4040",  "U.S. Auto Force — Parts",                 "Inventory-Parts & Supplies", 659.96),
    ("04/13/2026", "",      "ADP Wage Pay — Payroll",                  "Payroll-Wages",             1976.32),
    ("04/13/2026", "",      "ADP Tax Deposit",                         "Payroll-Payroll Taxes",      607.68),
    ("04/13/2026", "",      "ADP 401k Contribution",                   "Payroll-401k",                30.00),
    ("04/13/2026", "",      "AT&T Phone",                              "Utilities-Phone",            244.89),
    ("04/13/2026", "4051",  "Tire Outlet — Tire Purchase",             "Inventory-Tires",           1816.00),
    ("04/13/2026", "4061",  "Tire Outlet — Tire Purchase",             "Inventory-Tires",           1216.00),
    ("04/13/2026", "4072",  "Basil Hosmer — Rent",                     "Rent",                      3400.00),
    ("04/13/2026", "4073",  "Tire Outlet — Tire Purchase",             "Inventory-Tires",           1080.00),
    ("04/13/2026", "4074",  "U.S. Auto Force — Parts",                 "Inventory-Parts & Supplies", 120.12),
    ("04/13/2026", "4075",  "GW — Labor",                              "Labor/Subcontractors",       110.00),
    ("04/14/2026", "",      "ADP Pay-By-Pay (Workers Comp)",           "Payroll-Workers Comp",        86.75),
    ("04/14/2026", "4076",  "Tire Outlet — Tire Purchase",             "Inventory-Tires",            460.00),
    ("04/14/2026", "4077",  "GW — Labor",                              "Labor/Subcontractors",       150.00),
    ("04/15/2026", "",      "ADP Wage Pay — Payroll",                  "Payroll-Wages",             1976.32),
    ("04/16/2026", "",      "TAPCO Insurance — Business Insurance",    "Insurance-Business",        1138.82),
    ("04/16/2026", "",      "TimePaymentcorp — Equipment Lease",       "Equipment Lease",            741.95),
    ("04/16/2026", "4078",  "Tire Parts — Parts Purchase",             "Inventory-Parts & Supplies", 152.31),
    ("04/17/2026", "",      "ADP Payroll Processing Fee",              "Payroll-Processing Fees",    134.50),
    ("04/17/2026", "",      "Gladly Coffee",                           "Meals/Entertainment",         27.32),
    ("04/20/2026", "",      "ADP Wage Pay — Payroll",                  "Payroll-Wages",             1976.32),
    ("04/20/2026", "",      "ADP Tax Deposit",                         "Payroll-Payroll Taxes",      607.68),
    ("04/20/2026", "4079",  "Tire Co — Tire Purchase",                 "Inventory-Tires",            771.70),
    ("04/21/2026", "",      "ADP 401k Contribution",                   "Payroll-401k",                30.00),
    ("04/21/2026", "",      "ADP Pay-By-Pay (Workers Comp)",           "Payroll-Workers Comp",        86.75),
    ("04/21/2026", "",      "ADP Tax Deposit",                         "Payroll-Payroll Taxes",      607.68),
    ("04/22/2026", "4080",  "U.S. Auto Force — Parts",                 "Inventory-Parts & Supplies", 153.76),
    ("04/22/2026", "4081",  "Tireco — Tire Purchase",                  "Inventory-Tires",            318.80),
    ("04/23/2026", "4082",  "Tire Outlet — Tire Purchase",             "Inventory-Tires",            400.00),
    ("04/24/2026", "",      "ADP Payroll Processing Fee",              "Payroll-Processing Fees",    134.50),
    ("04/27/2026", "",      "ADP Wage Pay — Payroll",                  "Payroll-Wages",             1976.32),
    ("04/27/2026", "",      "ADP Tax Deposit",                         "Payroll-Payroll Taxes",      607.68),
    ("04/27/2026", "",      "Gladly Coffee",                           "Meals/Entertainment",          5.39),
    ("04/27/2026", "4083",  "Tire Network — Tire Purchase",            "Inventory-Tires",            120.00),
    ("04/28/2026", "",      "ADP 401k Contribution",                   "Payroll-401k",                30.00),
    ("04/28/2026", "",      "ADP Pay-By-Pay (Workers Comp)",           "Payroll-Workers Comp",        86.75),
    ("04/28/2026", "",      "Snapdragon Stadium",                      "Meals/Entertainment",         35.56),
    ("04/28/2026", "4085",  "Tire Network — Tire Purchase",            "Inventory-Tires",            228.00),
    ("04/29/2026", "",      "SANDAG FasTrak — Toll",                   "Transportation",               2.50),
    ("04/29/2026", "",      "APG Wholesale — Parts",                   "Inventory-Parts & Supplies", 415.80),
    ("04/29/2026", "",      "BMW El Cajon — Vehicle Repair",           "Vehicle-Repairs",            133.52),
    ("04/29/2026", "4071",  "Bureau of Auto Repair — License",         "Licenses/Permits",           250.00),
    ("04/29/2026", "4087",  "Tire Co — Tire Purchase",                 "Inventory-Tires",            276.84),
    ("04/30/2026", "4088",  "U.S. Auto Force — Parts",                 "Inventory-Parts & Supplies", 786.00),
    ("04/30/2026", "4092",  "Owner Draw — Francisco Reyes",            "Owner Draw",                 700.00),
]

exp_start = 3
for i, (dt, chk, desc, cat, amt) in enumerate(expenses):
    r = exp_start + i
    bg = GRAY if i % 2 == 0 else WHITE
    ws3.cell(row=r, column=1).value = dt
    ws3.cell(row=r, column=1).font = Font(name="Arial", size=10)
    ws3.cell(row=r, column=1).fill = hdr_fill(bg)
    ws3.cell(row=r, column=2).value = chk
    ws3.cell(row=r, column=2).font = Font(name="Arial", size=10)
    ws3.cell(row=r, column=2).fill = hdr_fill(bg)
    ws3.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    ws3.cell(row=r, column=3).value = desc
    ws3.cell(row=r, column=3).font = Font(name="Arial", size=10)
    ws3.cell(row=r, column=3).fill = hdr_fill(bg)
    ws3.cell(row=r, column=4).value = cat
    ws3.cell(row=r, column=4).font = Font(name="Arial", size=10)
    ws3.cell(row=r, column=4).fill = hdr_fill(bg)
    ws3.cell(row=r, column=5).value = amt
    ws3.cell(row=r, column=5).font = Font(name="Arial", size=10,
                                          color="0000FF")
    ws3.cell(row=r, column=5).number_format = CURR_FMT
    ws3.cell(row=r, column=5).fill = hdr_fill(bg)
    ws3.cell(row=r, column=5).alignment = Alignment(horizontal="right")
    ws3.row_dimensions[r].height = 15

exp_total_r = exp_start + len(expenses)
for col in range(1, 6):
    ws3.cell(row=exp_total_r, column=col).fill = hdr_fill(LT_BLUE)
ws3.cell(row=exp_total_r, column=1).value = "TOTAL"
ws3.cell(row=exp_total_r, column=1).font = Font(name="Arial", bold=True, size=10)
ws3.cell(row=exp_total_r, column=5).value = f"=SUM(E{exp_start}:E{exp_total_r-1})"
ws3.cell(row=exp_total_r, column=5).font = Font(name="Arial", bold=True, size=10)
ws3.cell(row=exp_total_r, column=5).number_format = CURR_FMT
ws3.cell(row=exp_total_r, column=5).fill = hdr_fill(LT_BLUE)
ws3.cell(row=exp_total_r, column=5).alignment = Alignment(horizontal="right")
ws3.row_dimensions[exp_total_r].height = 16

# Category breakdown (cols I-J)
ws3.cell(row=2, column=9).value = "Category"
ws3.cell(row=2, column=9).font = Font(name="Arial", bold=True, size=10,
                                       color="FFFFFF")
ws3.cell(row=2, column=9).fill = hdr_fill(NAVY)
ws3.cell(row=2, column=10).value = "Total"
ws3.cell(row=2, column=10).font = Font(name="Arial", bold=True, size=10,
                                        color="FFFFFF")
ws3.cell(row=2, column=10).fill = hdr_fill(NAVY)

exp_categories = [
    "Inventory-Tires",
    "Inventory-Parts & Supplies",
    "Payroll-Wages",
    "Payroll-Payroll Taxes",
    "Payroll-401k",
    "Payroll-Workers Comp",
    "Payroll-Processing Fees",
    "Rent",
    "Insurance-Business",
    "Equipment Lease",
    "Vehicle-Loan",
    "Vehicle-Repairs",
    "Utilities-Phone",
    "Merchant Processing",
    "Bank Fees",
    "Labor/Subcontractors",
    "Licenses/Permits",
    "Transportation",
    "Meals/Entertainment",
    "Miscellaneous",
    "Owner Draw",
    "Credit Card Payment",
]
for j, cat in enumerate(exp_categories):
    br = 3 + j
    ws3.cell(row=br, column=9).value = cat
    ws3.cell(row=br, column=9).font = Font(name="Arial", size=10)
    ws3.cell(row=br, column=9).fill = hdr_fill(GRAY if j % 2 == 0 else WHITE)
    formula = (f'=SUMIF($D${exp_start}:$D${exp_total_r-1},'
               f'I{br},$E${exp_start}:$E${exp_total_r-1})')
    ws3.cell(row=br, column=10).value = formula
    ws3.cell(row=br, column=10).font = Font(name="Arial", size=10)
    ws3.cell(row=br, column=10).number_format = CURR_FMT
    ws3.cell(row=br, column=10).fill = hdr_fill(GRAY if j % 2 == 0 else WHITE)
    ws3.cell(row=br, column=10).alignment = Alignment(horizontal="right")

exp_gt_r = 3 + len(exp_categories)
ws3.cell(row=exp_gt_r, column=9).value = "GRAND TOTAL"
ws3.cell(row=exp_gt_r, column=9).font = Font(name="Arial", bold=True, size=10)
ws3.cell(row=exp_gt_r, column=9).fill = hdr_fill(LT_BLUE)
ws3.cell(row=exp_gt_r, column=10).value = f"=SUM(J3:J{exp_gt_r-1})"
ws3.cell(row=exp_gt_r, column=10).font = Font(name="Arial", bold=True, size=10)
ws3.cell(row=exp_gt_r, column=10).number_format = CURR_FMT
ws3.cell(row=exp_gt_r, column=10).fill = hdr_fill(LT_BLUE)
ws3.cell(row=exp_gt_r, column=10).alignment = Alignment(horizontal="right")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – Credit Cards
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Credit Cards")
ws4.sheet_view.showGridLines = False

set_col_widths(ws4, {"A": 12, "B": 12, "C": 16, "D": 36, "E": 28,
                      "F": 6, "G": 16, "H": 4, "I": 22, "J": 14,
                      "K": 14, "L": 14})

ws4.merge_cells("A1:G1")
ws4["A1"].value = "REYES TIRES, INC — CREDIT CARD TRANSACTIONS — APRIL 2026"
ws4["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
ws4["A1"].fill = hdr_fill(NAVY)
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 26

cc_hdrs = ["Post Date", "Tx Date", "Account", "Description",
            "Category", "Flag", "Amount"]
for i, lbl in enumerate(cc_hdrs, 1):
    c = ws4.cell(row=2, column=i)
    c.value = lbl
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = hdr_fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[2].height = 16

cc_transactions = [
    # post, tx, account, description, category, flag, amount, is_personal
    ("04/01/2026", "04/01/2026", "...0214", "SDFC Team Gym",             "Personal - Gym",                        "⚠️", 46.08,   True),
    ("04/03/2026", "04/02/2026", "...0214", "Apple.com",                 "Personal - Subscriptions",              "⚠️",  9.99,   True),
    ("04/05/2026", "04/04/2026", "...0214", "SDFC Team Gym",             "Personal - Gym",                        "⚠️", 503.79,  True),
    ("04/06/2026", "04/05/2026", "...0214", "Amazon Digital",            "Personal - Subscriptions",              "⚠️", 12.99,   True),
    ("04/07/2026", "04/06/2026", "...0214", "Amazon Digital",            "Personal - Subscriptions",              "⚠️",  2.99,   True),
    ("04/08/2026", "04/07/2026", "...0214", "Amazon Prime",              "Personal - Subscriptions",              "⚠️", 16.15,   True),
    ("04/09/2026", "04/08/2026", "...0214", "ESPN+",                     "Personal - Subscriptions",              "⚠️", 19.99,   True),
    ("04/10/2026", "04/09/2026", "...0214", "Apple.com",                 "Personal - Subscriptions",              "⚠️", 22.98,   True),
    ("04/11/2026", "04/10/2026", "...0214", "Amazon Digital",            "Personal - Subscriptions",              "⚠️", 16.99,   True),
    ("04/12/2026", "04/11/2026", "...0214", "Amazon Digital",            "Personal - Subscriptions",              "⚠️",  2.99,   True),
    ("04/14/2026", "04/13/2026", "...0214", "SDFC Team Gym",             "Personal - Gym",                        "⚠️", 432.33,  True),
    ("04/15/2026", "04/14/2026", "...0214", "Amazon Digital",            "Personal - Subscriptions",              "⚠️", 19.99,   True),
    ("04/18/2026", "04/17/2026", "...0214", "Cox Communications",        "Utilities - Internet",                  "",   110.17,  False),
    ("04/20/2026", "04/19/2026", "...0214", "Infinity Kat",              "Miscellaneous (Review)",                "",   400.60,  False),
    ("04/22/2026", "04/21/2026", "...0214", "A to Z Tax Firm",           "Professional Services - Tax Prep",      "",   420.00,  False),
    ("04/30/2026", "04/30/2026", "...0214", "Finance Charge",            "Finance Charges",                       "",    63.07,  False),
    ("04/30/2026", "04/30/2026", "...0539", "Finance Charge",            "Finance Charges",                       "",    51.72,  False),
    ("04/15/2026", "04/14/2026", "...6514", "ADT Security Services",     "Security System",                       "",    98.08,  False),
    ("04/30/2026", "04/30/2026", "...6514", "Finance Charge",            "Finance Charges",                       "",    96.97,  False),
]

cc_start = 3
for i, (post, tx, acct, desc, cat, flag, amt, is_personal) in enumerate(cc_transactions):
    r = cc_start + i
    bg = YELLOW if is_personal else (GRAY if i % 2 == 0 else WHITE)
    vals = [post, tx, acct, desc, cat, flag, amt]
    for col_idx, val in enumerate(vals, 1):
        cell = ws4.cell(row=r, column=col_idx)
        cell.fill = hdr_fill(bg)
        if col_idx == 7:
            cell.value = val
            cell.font = Font(name="Arial", size=10,
                             color="0000FF" if not is_personal else "000000",
                             bold=is_personal)
            cell.number_format = CURR_FMT
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.value = val
            cell.font = Font(name="Arial", size=10,
                             bold=is_personal)
            if col_idx == 6:
                cell.alignment = Alignment(horizontal="center")
    ws4.row_dimensions[r].height = 15

cc_total_r = cc_start + len(cc_transactions)
for col in range(1, 8):
    ws4.cell(row=cc_total_r, column=col).fill = hdr_fill(LT_BLUE)
ws4.cell(row=cc_total_r, column=1).value = "TOTAL"
ws4.cell(row=cc_total_r, column=1).font = Font(name="Arial", bold=True, size=10)
ws4.cell(row=cc_total_r, column=7).value = f"=SUM(G{cc_start}:G{cc_total_r-1})"
ws4.cell(row=cc_total_r, column=7).font = Font(name="Arial", bold=True, size=10)
ws4.cell(row=cc_total_r, column=7).number_format = CURR_FMT
ws4.cell(row=cc_total_r, column=7).fill = hdr_fill(LT_BLUE)
ws4.cell(row=cc_total_r, column=7).alignment = Alignment(horizontal="right")
ws4.row_dimensions[cc_total_r].height = 16

# Personal charges note
note_r = cc_total_r + 2
ws4.merge_cells(f"A{note_r}:G{note_r}")
ws4[f"A{note_r}"].value = ("⚠️ Yellow rows = Personal charges on business cards. "
                            "Total personal charges: ~$1,107.26. "
                            "Require owner reclassification / reimbursement.")
ws4[f"A{note_r}"].font = Font(name="Arial", size=9, bold=True, color="FF0000")
ws4[f"A{note_r}"].fill = hdr_fill("FFF3CD")
ws4[f"A{note_r}"].alignment = Alignment(horizontal="left", vertical="center",
                                         wrap_text=True)
ws4.row_dimensions[note_r].height = 28

# Card balance table (cols I-L)
ws4.cell(row=2, column=9).value = "Card"
ws4.cell(row=2, column=10).value = "Limit"
ws4.cell(row=2, column=11).value = "Balance"
ws4.cell(row=2, column=12).value = "Available"
for col in range(9, 13):
    ws4.cell(row=2, column=col).font = Font(name="Arial", bold=True, size=10,
                                             color="FFFFFF")
    ws4.cell(row=2, column=col).fill = hdr_fill(NAVY)
    ws4.cell(row=2, column=col).alignment = Alignment(horizontal="center")

card_balances = [
    ("...0214 (F. Reyes)", 4000.00, 3807.94),
    ("...0539",           38000.00, 2990.95),
    ("...6514",           47000.00, 4546.94),
]
for j, (card, limit, bal) in enumerate(card_balances):
    br = 3 + j
    bg = GRAY if j % 2 == 0 else WHITE
    ws4.cell(row=br, column=9).value = card
    ws4.cell(row=br, column=9).font = Font(name="Arial", size=10)
    ws4.cell(row=br, column=9).fill = hdr_fill(bg)
    ws4.cell(row=br, column=10).value = limit
    ws4.cell(row=br, column=10).font = Font(name="Arial", size=10,
                                             color="0000FF")
    ws4.cell(row=br, column=10).number_format = CURR_FMT
    ws4.cell(row=br, column=10).fill = hdr_fill(bg)
    ws4.cell(row=br, column=10).alignment = Alignment(horizontal="right")
    ws4.cell(row=br, column=11).value = bal
    ws4.cell(row=br, column=11).font = Font(name="Arial", size=10,
                                             color="0000FF")
    ws4.cell(row=br, column=11).number_format = CURR_FMT
    ws4.cell(row=br, column=11).fill = hdr_fill(bg)
    ws4.cell(row=br, column=11).alignment = Alignment(horizontal="right")
    ws4.cell(row=br, column=12).value = f"=J{br}-K{br}"
    ws4.cell(row=br, column=12).font = Font(name="Arial", size=10)
    ws4.cell(row=br, column=12).number_format = CURR_FMT
    ws4.cell(row=br, column=12).fill = hdr_fill(bg)
    ws4.cell(row=br, column=12).alignment = Alignment(horizontal="right")
    ws4.row_dimensions[br].height = 15

# Total row for card balances
bal_total_r = 3 + len(card_balances)
for col in range(9, 13):
    ws4.cell(row=bal_total_r, column=col).fill = hdr_fill(LT_BLUE)
ws4.cell(row=bal_total_r, column=9).value = "TOTAL"
ws4.cell(row=bal_total_r, column=9).font = Font(name="Arial", bold=True, size=10)
for col_idx, col_ltr in [(10, "J"), (11, "K"), (12, "L")]:
    ws4.cell(row=bal_total_r, column=col_idx).value = (
        f"=SUM({col_ltr}3:{col_ltr}{bal_total_r-1})")
    ws4.cell(row=bal_total_r, column=col_idx).font = Font(name="Arial",
                                                           bold=True, size=10)
    ws4.cell(row=bal_total_r, column=col_idx).number_format = CURR_FMT
    ws4.cell(row=bal_total_r, column=col_idx).fill = hdr_fill(LT_BLUE)
    ws4.cell(row=bal_total_r, column=col_idx).alignment = Alignment(
        horizontal="right")
ws4.row_dimensions[bal_total_r].height = 16

# ─── Save ──────────────────────────────────────────────────────────────────
OUT = "/Users/reyesinc-op1/Documents/roveloinc.com/Reyes_Tires_Inc_April2026_Bookkeeping.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
